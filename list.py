#!/usr/bin/python3

import os
import subprocess
from openpyxl import Workbook

pwd = str(os.getcwd())
len_pwd = len(pwd)
if len_pwd == 3:  # If pwd is a drive and not a directory
    delete = 3  # Deleting drive letter and "\", for example "E:\"
else:  # If inside a directory
    delete = len_pwd+1
entries = []
for p, d, f in os.walk(pwd):  # p=>same as pwd, d=>directories, f=>files
    for item in d:
        temp = str(os.path.join(p, item))  # abs. path
        entries.append(temp[delete:])
    for item in f:
        temp = str(os.path.join(p, item))
        entries.append(temp[delete:])
entries.sort()    

wb = Workbook()
ws = wb.active
ws.title = "Tree"
ws.cell(row=2, column=1, value="Sl. No.")
ws.cell(row=2, column=2, value="Entry")
for i in range(len(entries)):
    ws.cell(row=i+3, column=1, value=i+1)
    ws.cell(row=i+3, column=2, value=entries[i])

# For optimal column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

ws.merge_cells("A1:B1")  # Merging cells after setting optimal column width. Otherwise .column_letter wont return anything.
ws.cell(row=1, column=1, value="Entries in the tree")

wb.save("SyncInfo.xlsx")
os.system("notify-send \"SyncInfo.xlsx, created!\"")
subprocess.run(["echo", "-e", "\a"], check=True)
