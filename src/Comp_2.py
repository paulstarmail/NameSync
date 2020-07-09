import os, openpyxl, winsound, ctypes
from openpyxl.styles import Alignment

pwd = str(os.getcwd())
delete = len(pwd) + 1  # 1 for deleting "\"
entries = []
for r, d, f in os.walk(pwd):  # r=>root, d=>directories, f=>files
    for item in d:
        temp = str(os.path.join(r, item))
        entries.append(temp[delete:])  # Making the abs. path to relative path
    for item in f:
        temp = str(os.path.join(r, item))
        entries.append(temp[delete:])
entries.sort()    

#### Comp_2 ####

wb = openpyxl.load_workbook("SyncInfo.xlsx")
wb.create_sheet("Comp_2", 0) # insert at first position
ws = wb.active
ws.cell(row=2, column=1, value="Sl. No.")
ws.cell(row=2, column=2, value="Entry")
for i in range(len(entries)):
    ws.cell(row=i+3, column=1, value=i+1)
    ws.cell(row=i+3, column=2, value=entries[i])

# For optimal column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

ws.merge_cells("A1:B1")  # Merging cells after setting optimal column width. Otherwise .column_letter wont return anything.
ws.cell(row=1, column=1, value="Entries in Comp_2")

#### To Comp_1 ####

last_row_comp_1 = 3
last_row_comp_2 = 3
found = 0
entries = []  # List cleared

ws = wb["Comp_1"]
while True: # To find max no of rows
    if ws.cell(row=last_row_comp_1, column=2).value is None:
        break
    last_row_comp_1 += 1

ws = wb["Comp_2"]
while True: # To find max no of rows
    if ws.cell(row=last_row_comp_2, column=2).value is None:
        break
    last_row_comp_2 += 1
    
for comp_2_i in range(last_row_comp_2):
    ws = wb["Comp_2"]
    comp_2_entry = ws.cell(row=comp_2_i+3, column=2).value
    

    ws = wb["Comp_1"]
    for comp_1_i in range(last_row_comp_1):
        comp_1_entry = ws.cell(row=comp_1_i+3, column=2).value
        
        if str(comp_2_entry) == str(comp_1_entry):
            found = 1
            break
    if found == 0:
        entries.append(comp_2_entry)
    else:    
        found = 0
entries.sort()

wb.create_sheet("To Comp_1", 0) # insert at first position
ws = wb.active
ws.cell(row=2, column=1, value="Sl. No.")
ws.cell(row=2, column=2, value="Entry")

for i in range(len(entries)):
    ws.cell(row=i+3, column=1, value=i+1)
    ws.cell(row=i+3, column=2, value=entries[i])

# For optimal column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

ws.merge_cells("A1:B1")  # Merging cells after setting optimal column width. Otherwise .column_letter wont return anything.
ws.cell(row=1, column=1, value="Entries to be copied from Comp_2 to Comp_1")
ws['A1'].alignment = Alignment(wrap_text=True)
ws.row_dimensions[1].height = 28

### To Comp_2 ####
found = 0
entries = []  # List cleared

for comp_1_i in range(last_row_comp_1):
    ws = wb["Comp_1"]
    comp_1_entry = ws.cell(row=comp_1_i+3, column=2).value
    

    ws = wb["Comp_2"]
    for comp_2_i in range(last_row_comp_2):
        comp_2_entry = ws.cell(row=comp_2_i+3, column=2).value
        
        if str(comp_1_entry) == str(comp_2_entry):
            found = 1
            break
    if found == 0:
        entries.append(comp_1_entry)
    else:    
        found = 0
entries.sort()

wb.create_sheet("To Comp_2", 0) # insert at first position
ws = wb.active
ws.cell(row=2, column=1, value="Sl. No.")
ws.cell(row=2, column=2, value="Entry")

for i in range(len(entries)):
    ws.cell(row=i+3, column=1, value=i+1)
    ws.cell(row=i+3, column=2, value=entries[i])

# For optimal column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

ws.merge_cells("A1:B1")  # Merging cells after setting optimal column width. Otherwise .column_letter wont return anything.
ws.cell(row=1, column=1, value="Entries to be copied from Comp_1 to Comp_2")
ws['A1'].alignment = Alignment(wrap_text=True)
ws.row_dimensions[1].height = 28

#### Saving and Notifying ####
wb.save("SyncInfo.xlsx")
winsound.Beep(2000, 1000)
ctypes.windll.user32.MessageBoxW(0, "SyncInfo modification complete !", "SyncInfo Modified", 1)
