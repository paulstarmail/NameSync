import os, openpyxl, winsound, ctypes
from openpyxl.styles import Alignment

pwd = str(os.getcwd())
delete = len(pwd) + 1  # 1 for deleting "\"
entries = []
for r, d, f in os.walk(pwd):  # r=>root, d=>directories, f=>files
    for item in d:
        temp = str(os.path.join(r, item))
        entries.append(temp[delete:])  # Making the abs. path to relative path
    for item in f:
        temp = str(os.path.join(r, item))
        entries.append(temp[delete:])
entries.sort()    

#### Source ####

wb = openpyxl.load_workbook("SyncInfo.xlsx")
wb.create_sheet("Source", 0) # insert at first position
ws = wb.active
ws.cell(row=2, column=1, value="Sl. No.")
ws.cell(row=2, column=2, value="Entry")
for i in range(len(entries)):
    ws.cell(row=i+3, column=1, value=i+1)
    ws.cell(row=i+3, column=2, value=entries[i])

# For optimal column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

ws.merge_cells("A1:B1")  # Merging cells after setting optimal column width. Otherwise .column_letter wont return anything.
ws.cell(row=1, column=1, value="Entries in Source")

#### To Sink ####

last_row_sink = 3
last_row_source = 3
found = 0
entries = []  # List cleared

ws = wb["Sink"]
while True:
    if ws.cell(row=last_row_sink, column=2).value is None:
        break
    last_row_sink += 1

ws = wb["Source"]
while True:
    if ws.cell(row=last_row_source, column=2).value is None:
        break
    last_row_source += 1
    
for src_i in range(last_row_source):
    ws = wb["Source"]
    src_entry = ws.cell(row=src_i+3, column=2).value
    

    ws = wb["Sink"]
    for sink_i in range(last_row_sink):
        sink_entry = ws.cell(row=sink_i+3, column=2).value
        
        if str(src_entry) == str(sink_entry):
            found = 1
            break
    if found == 0:
        entries.append(src_entry)
    else:    
        found = 0
entries.sort()

wb.create_sheet("To Sink", 0) # insert at first position
ws = wb.active
ws.cell(row=2, column=1, value="Sl. No.")
ws.cell(row=2, column=2, value="Entry")

for i in range(len(entries)):
    ws.cell(row=i+3, column=1, value=i+1)
    ws.cell(row=i+3, column=2, value=entries[i])

# For optimal column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

ws.merge_cells("A1:B1")  # Merging cells after setting optimal column width. Otherwise .column_letter wont return anything.
ws.cell(row=1, column=1, value="Entries to be copied from Source to Sink")
ws['A1'].alignment = Alignment(wrap_text=True)
ws.row_dimensions[1].height = 28

wb.save("SyncInfo.xlsx")
winsound.Beep(2000, 1000)
ctypes.windll.user32.MessageBoxW(0, "SyncInfo modification complete !", "SyncInfo Modified", 1)
