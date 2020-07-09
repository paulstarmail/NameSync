import os, winsound, ctypes
from openpyxl import Workbook

pwd = str(os.getcwd())
delete = len(pwd) + 1  # 1 for deleting "\"
entries = []
for r, d, f in os.walk(pwd):  # r=>root, d=>directories, f=>files
    for item in d:
        temp = str(os.path.join(r, item))
        entries.append(temp[delete:])  # Making the abs. path to relative path
    for item in f:
        temp = str(os.path.join(r, item))
        entries.append(temp[delete:])
entries.sort()    

wb = Workbook()
ws = wb.active
ws.title = "Comp_1"
ws.cell(row=2, column=1, value="Sl. No.")
ws.cell(row=2, column=2, value="Entry")
for i in range(len(entries)):
    ws.cell(row=i+3, column=1, value=i+1)
    ws.cell(row=i+3, column=2, value=entries[i])

# For optimal column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

ws.merge_cells("A1:B1")  # Merging cells after setting optimal column width. Otherwise .column_letter wont return anything.
ws.cell(row=1, column=1, value="Entries in Comp_1")

wb.save("SyncInfo.xlsx")
winsound.Beep(2000, 1000)
ctypes.windll.user32.MessageBoxW(0, "SyncInfo creation complete !", "SyncInfo Created", 1)
